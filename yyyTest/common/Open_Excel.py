import re, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, PatternFill
from common.up_data import up_excel


# 数据预处理，去掉所有空格
def update_excel(excel_path, sheet_name):
    for i in range(2):
        try:
            data = load_workbook(excel_path)
            sheet = data[sheet_name]
        except Exception as e:
            print('测试用例文件打开错误', e)
        else:
            max_rows = sheet.max_row
            max_cols = sheet.max_column
            for x in range(1, max_rows + 1):
                for y in range(1, max_cols + 1):
                    a = sheet.cell(row=x, column=y).value
                    if isinstance(a, str):
                        s = re.sub('/s', '', a)
                        if s == '':
                            sheet.cell(x, y, '')
                        else:
                            pass
                    else:
                        pass
            data.save(excel_path)


# 读取Excel公共方法:列
def read_excel(excel_path, sheet_name, col):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误', e)
    else:
        newdata = []
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=col).value is None:
                newdata.append("")
            else:
                newdata.append(sheet.cell(row=r, column=col).value)
        return newdata


# 读取Excel公共方法:行
def read_excel_row(excel_path, sheet_name, row_name, field=[], valu=[]):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误', e)
    else:
        maxr = sheet.max_row
        maxl = sheet.max_column
        r = row_name
        dict = {}
        data_dict = []
        for i in range(r, r + 1):
            for j in range(1, maxl - 1):
                title = sheet.cell(row=1, column=j).value
                value = sheet.cell(row=i, column=j).value
                new_value = up_excel(value)  # 获取自定义参数
                dict[title] = new_value
            info = copy.deepcopy(dict)
            # 修改自定的用例参数
            if not field and not valu:
                pass
            else:
                for i, j in zip(field, valu):
                    info[i] = j
            data_dict.append(info)
        return data_dict


# 写入Excel公共方法
def write_excel(excel_path, sheet_name, result, row_name):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误', e)
    else:
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        i = row_name
        sheet.cell(i, max_cols - 1, result)
        ippass = sheet.cell(row=i, column=max_cols)
        actual = sheet.cell(row=i, column=max_cols - 1)
        expected = sheet.cell(row=i, column=max_cols - 2)
        if actual.value == expected.value:
            sheet.cell(i, max_cols, 'PASS')
            ippass.font = Font(color='00FF00')
            ippass.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        else:
            sheet.cell(i, max_cols, 'FAIL')
            ippass.font = Font(color='CC3300')
            ippass.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        data.save(excel_path)
        data.close()
