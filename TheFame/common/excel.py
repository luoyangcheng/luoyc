from openpyxl import load_workbook
from openpyxl.styles import Font, colors


# 读取Excel公共方法
def read_excel(excel_path, sheet_name, col):
    data = load_workbook(excel_path)
    sheet = data[sheet_name]
    newdata = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=col).value is None:
            newdata.append("")
        else:
            newdata.append(sheet.cell(row=r, column=col).value)

    return newdata


# 写入Excel公共方法
def write_excel(excel_path, sheet_name, Result):
    green = Font(color=colors.GREEN)
    red = Font(color=colors.RED)
    data = load_workbook(excel_path)
    sheet = data[sheet_name]
    rows = sheet.max_row
    cols = sheet.max_column
    l_ = []
    for i in range(2, rows + 1):
        l_.append(i)

    for i, j in zip(l_, Result):
        sheet.cell(i, cols, j)
        actual = sheet.cell(row=i, column=cols)
        expected = sheet.cell(row=i, column=cols-1)
        if actual.value == expected.value:
            resu.font = green
        else:
            resu.font = red
    data.save(excel_path)
