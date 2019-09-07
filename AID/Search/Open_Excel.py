import re
from openpyxl import load_workbook
from openpyxl.styles import Font, colors


# 数据预处理，去掉所有空格
def update_excel(excel_path, sheet_name):
    for i in range(2):
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        for x in range(1, max_rows + 1):
            for y in range(1, max_cols + 1):
                a = sheet.cell(row=x, column=y).value
                if type(a) == type(""):
                    s = re.sub('\s', '', a)
                    if s == '':
                        sheet.cell(x, y, '')
                    else:
                        pass
                else:
                    pass
        data.save(excel_path)

# 读取Excel公共方法
def read_excel(excel_path, sheet_name, col):
    update_excel(excel_path, sheet_name)
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误', e)
    else:
        newdata = []
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=col).value is None:
                newdata.append("")
            else:
                newdata.append(sheet.cell(row=r, column=col).value)
        return newdata


# 写入Excel公共方法
def write_excel(excel_path, sheet_name, Result):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误')
    else:
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        for i in range(2, max_rows + 1):
            sheet.cell(i, max_cols - 1, Result[i - 2])
            ippass = sheet.cell(row=i, column=max_cols)
            actual = sheet.cell(row=i, column=max_cols - 1)
            expected = sheet.cell(row=i, column=max_cols - 2)
            if actual.value == expected.value:
                sheet.cell(i, max_cols, 'PASS')
                ippass.font = Font(color=colors.GREEN)
            else:
                sheet.cell(i, max_cols, 'FAIL')
                ippass.font = Font(color=colors.RED)
        data.save(excel_path)
        data.close()