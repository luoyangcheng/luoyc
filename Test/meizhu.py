# -*- coding: utf-8 -*-

from selenium import webdriver
from Meizhu._def import waitid,waitxp
from openpyxl import Workbook
from openpyxl import load_workbook


def read_excel():
    data=load_workbook('C:/Users\Administrator\luoyc\Meizhu\meizhu_testcase.xlsx')
    sheet=data["美住登录"]
    username = []
    for r in range(2, sheet.max_row + 1):
        username.append(sheet.cell(row=r, column=1).value)
    passwd = []
    for r in range(2, sheet.max_row + 1):
        passwd.append(sheet.cell(row=r, column=2).value)
    tip = []
    for r in range(2, sheet.max_row + 1):
        tip.append(sheet.cell(row=r, column=3).value)
    return username , passwd,tip

def browser():
    username,passwd,tip=read_excel()
    url = "http://192.168.3.19:8090/login.html"
    browser = webdriver.Firefox()
    result=""
    rr=[]
    for x, y, z in zip(username, passwd, tip):
        browser.get(url)
        waitid(browser,"requestUsername")
        browser.find_element_by_id("requestUsername").send_keys(x)
        waitid(browser, "requestPassword")
        browser.find_element_by_id("requestPassword").send_keys(y)
        waitid(browser, "requestSubmit")
        browser.find_element_by_id("requestSubmit").click()
        str = browser.find_element_by_id("login-tip").text
        if str == z:
            print("测试通过")
            result = "PASS"
            rr.append(result)
        else:
            print("\033[1;31m !!!!ERRO!!!! \033[0m!", x, y,str,z)
            result = "FAIL"
            rr.append(result)
    return rr

def write_excel():
    rr=browser()
    data = load_workbook('C:/Users\Administrator\luoyc\Meizhu\meizhu_testcase.xlsx')
    sheet = data["美住登录"]
    rows = sheet.max_row
    l_ = []
    for i in range(2, rows + 1):
        l_.append(i)
    print(l_)

    for i, j in zip(l_, rr):
        sheet.cell(i, 4, j)
    data.save('C:/Users\Administrator\luoyc\Meizhu\meizhu_testcase.xlsx')


if __name__ == "__main__":
    write_excel()
