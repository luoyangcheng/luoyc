from openpyxl import load_workbook
from termcolor import *
from openpyxl.styles import Font, colors, Alignment

import xlwt
bold_itatic_24_font = Font(color=colors.RED)

data=load_workbook('F:/text.xlsx')
sheet=data["Sheet1"]
sheet['C3'].font = bold_itatic_24_font
rows=sheet.max_row
a = "FAIL"
sheet.cell(3,3,a)
data.save('F:/text.xlsx')



