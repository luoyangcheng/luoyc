# 递归函数，比如可以用来计算:1*2*3...*n = ?
from openpyxl import load_workbook
from openpyxl.styles import Font, colors


def read_excel(excel_path, sheet_name, col):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误', e)
    else:
        newdata = []
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=col).value is None:
                newdata.append("")
            else:
                newdata.append(sheet.cell(row=r, column=col).value)
        return newdata


def write_excel(excel_path, sheet_name, Result):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误')
    else:
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        print(max_cols)          
        for i in range(2, max_rows+1):
            sheet.cell(i, max_cols - 1, Result[i-2])
            ippass = sheet.cell(row=i, column=max_cols)
            actual = sheet.cell(row=i, column=max_cols - 1)
            expected = sheet.cell(row=i, column=max_cols - 2)
            if actual.value == expected.value:
                sheet.cell(i, max_cols, 'PASS')
                ippass.font = Font(color=colors.GREEN)
            else:
                sheet.cell(i, max_cols, 'FAIL')
                ippass.font = Font(color=colors.RED)
        data.save(excel_path)
        data.close()


if __name__ == '__main__':
    write_excel('D:/a.xlsx', 'Sheet1', ['结果', '结果'])
