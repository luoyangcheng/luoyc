import re, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, PatternFill


# 读取Excel公共方法:行
def read_excel_row(excel_path, sheet_name, row_name, field=[], valu=[]):
    try:
        data = load_workbook(excel_path)
        sheet = data[sheet_name]
    except Exception as e:
        print('测试用例文件打开错误', e)
    else:
        maxr = sheet.max_row
        maxl = sheet.max_column
        r = row_name
        dict = {}
        data_dict = []
        for i in range(r, r):
            for j in range(1, maxl - 1):
                title = sheet.cell(row=1, column=j).value
                value = sheet.cell(row=i, column=j).value
                new_value = up_excel(value)  # 获取自定义参数
                dict[title] = new_value
            info = copy.deepcopy(dict)
            # 修改自定的用例参数
            if not field and not valu:
                pass
            else:
                for i, j in zip(field, valu):
                    info[i] = j
            data_dict.append(info)
        return data_dict