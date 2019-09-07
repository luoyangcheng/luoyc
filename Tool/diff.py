from openpyxl import load_workbook


def read_excel(excel_path, sheet_name, col):
    data = load_workbook(excel_path)
    sheet = data[sheet_name]
    newdata = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=col).value is None:
            newdata.append("")
        else:
            newdata.append(sheet.cell(row=r, column=col).value)

    return newdata


def aa():
    excel_path = "D:/test.xlsx"
    a = read_excel(excel_path, 'Sheet1', 4)
    b = read_excel(excel_path, 'Sheet1', 5)
    s1 = set(a)
    s2 = set(b)
    # 以list1为主，list2中缺少的 [1]
    s = s1.difference(s2)
    print(s)

    # s1 s2中不同的地方
    s = s1.symmetric_difference(s2)
    print(s)
    # s1 s2中相同的地方
    s = s1.intersection(s2)
    print(s)
    # 合并两个set
    s = s1.union(s2)
    print(s)
    # s1是不是s2的子集
    s = s1.issubset(s2)
    print(s)
    # s1是不是s2的超集
    s = s1.issuperset(s2)
    print(s)


if __name__ == "__main__":
    aa()
