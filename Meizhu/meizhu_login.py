# 美住登录功能
from selenium import webdriver
from _def import waitid, waittext
from openpyxl import load_workbook
from termcolor import colored
from openpyxl.styles import Font, colors


class Login:
    def read_excel(self):
        data = load_workbook('E:/luoyc\Meizhu\meizhu_testcase.xlsx')
        sheet = data["美住登录"]
        username = []
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=1).value is None:
                username.append("")
            else:
                username.append(sheet.cell(row=r, column=1).value)

        passwd = []
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=2).value is None:
                passwd.append("")
            else:
                passwd.append(sheet.cell(row=r, column=2).value)

        tip = []
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=3).value is None:
                tip.append("")
            else:
                tip.append(sheet.cell(row=r, column=3).value)
        return username, passwd, tip

    def browser(self):
        username, passwd, tip = self.read_excel()
        url = "http://192.168.3.19:8090/login.html"
        browser = webdriver.Firefox()
        browser.implicitly_wait(3)
        Result = []
        for x, y, z in zip(username, passwd, tip):
            browser.get(url)
            waitid(browser, "requestUsername")
            browser.find_element_by_id("requestUsername").send_keys(x)
            waitid(browser, "requestPassword")
            browser.find_element_by_id("requestPassword").send_keys(y)
            waitid(browser, "requestSubmit")
            browser.find_element_by_id("requestSubmit").click()
            tip = waittext(browser, "login-tip")
            if tip == z:
                print(colored('测试通过', 'green'))
                result_text = "PASS"
                Result.append(result_text)
            else:
                print(colored('测试失败', 'red'))
                result_text = "FAIL"
                Result.append(result_text)
        return Result

    def write_excel(self):
        Result = self.browser()
        green = Font(color=colors.GREEN)
        red = Font(color=colors.RED)
        data = load_workbook('E:/luoyc\Meizhu\meizhu_testcase.xlsx')
        sheet = data["美住登录"]
        rows = sheet.max_row
        cols = sheet.max_column
        l_ = []
        for i in range(2, rows + 1):
            l_.append(i)

        for i, j in zip(l_, Result):
            sheet.cell(i, cols, j)
            resu = sheet.cell(row=i, column=cols)
            if resu.value == "PASS":
                resu.font = green
            else:
                resu.font = red
        data.save('E:/luoyc\Meizhu\meizhu_testcase.xlsx')


if __name__ == "__main__":
    Login().write_excel()
