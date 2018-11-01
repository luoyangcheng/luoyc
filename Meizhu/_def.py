# 显性等待时间公共函数
# 通过xpath id ...查找页面元素
# 一直查找5秒、每0.5秒查找一次
# 如果找不到，返回空字符串，比如登陆成功就没有提示等等

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Font, colors


def waitxp(browser, xpath):
    locator = (By.XPATH, xpath)
    try:
        WebDriverWait(browser, 5, 0.5).until(
            EC.visibility_of_element_located(locator))
    except Exception as e:
        print(e)


def waitid(browser, id):
    locator = (By.ID, id)
    try:
        WebDriverWait(browser, 5, 0.5).until(
            EC.visibility_of_element_located(locator))
    except Exception as e:
        print(e)


def waittext(browser, id):
    locator = (By.ID, id)
    try:
        WebDriverWait(browser, 5, 0.5).until(
            EC.visibility_of_element_located(locator))
        return browser.find_element_by_id(id).text
    except Exception as e:
        print(e)
        return ""


# 读取Excel公共方法
def read_excel(excel_path, sheet_name, col):
    data = load_workbook(excel_path)
    sheet = data[sheet_name]
    newdata = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=col).value is None:
            newdata.append("")
        else:
            newdata.append(sheet.cell(row=r, column=col).value)

    return newdata


# 写入Excel公共方法
def write_excel(excel_path, sheet_name, Result):
    green = Font(color=colors.GREEN)
    red = Font(color=colors.RED)
    data = load_workbook(excel_path)
    sheet = data[sheet_name]
    rows = sheet.max_row
    cols = sheet.max_column
    l_ = []
    for i in range(2, rows + 1):
        l_.append(i)

    for i, j in zip(l_, Result):
        sheet.cell(i, cols, j)
        resu = sheet.cell(row=i, column=cols)
        if resu.value == "PASS":
            resu.font = green
        else:
            resu.font = red
    data.save(excel_path)
