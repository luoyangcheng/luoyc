#
# 提交订单-间夜房
import sys
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
from Meizhu._def import waitid, waitxp, waitclass
import time
import xlrd


def read_excel():
    data = load_workbook('C:/Users\Administrator\luoyc\Meizhu\meizhu_testcase.xlsx')
    sheet = data["间夜房-提交订单"]
    orderusername = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=1).value is None:
            orderusername.append("")
        else:
            orderusername.append(sheet.cell(row=r, column=1).value)
    mobile = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=2).value is None:
            mobile.append("")
        else:
            mobile.append(sheet.cell(row=r, column=2).value)
    idcard = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=3).value is None:
            idcard.append("")
        else:
            idcard.append(sheet.cell(row=r, column=3).value)
    money = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=4).value is None:
            money.append("")
        else:
            money.append(sheet.cell(row=r, column=4).value)
    deposit = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=5).value is None:
            deposit.append("")
        else:
            deposit.append(sheet.cell(row=r, column=5).value)
    tip = []
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=6).value is None:
            tip.append("")
        else:
            tip.append(sheet.cell(row=r, column=6).value)

    return orderusername, mobile, idcard, money, deposit, tip


def login(browser):
    url = "http://192.168.3.19:8090/login.html"
    browser.get(url)
    browser.find_element_by_id("requestUsername").send_keys("18802094078")
    browser.find_element_by_id("requestPassword").send_keys("qq111111")
    browser.find_element_by_id("requestSubmit").click()


def booking(browser):
    orderusername, mobile, idcard, money, deposit, tip = read_excel()
    Result = []
    waitxp(browser, '//*[@id="orderListBody"]/tr[1]/td[4]/div')
    browser.find_element_by_xpath('//*[@id="orderListBody"]/tr[1]/td[4]/div').click()
    time.sleep(3)

    for i, j, k, x, y, t in zip(orderusername, mobile, idcard, money, deposit, tip):
        browser.find_element_by_xpath('//*[@id="addOrderRoom"]/table/tbody/tr[2]/td[1]/input').send_keys(i)
        browser.find_element_by_xpath('//*[@id="addOrderRoom"]/table/tbody/tr[2]/td[2]/div[2]/input').send_keys(j)
        browser.find_element_by_xpath('//*[@id="addOrderRoom"]/table/tbody/tr[2]/td[3]/input').send_keys(k)
        browser.find_element_by_xpath('//*[@id="addOrderReceive"]/tbody/tr/td[4]/div/input').send_keys(x)
        waitid(browser, "addReceipt")
        browser.find_element_by_id("addReceipt").click()
        waitxp(browser, '//*[@id="addOrderReceive"]/tbody/tr[2]/td[1]/div/a')
        browser.find_element_by_xpath('//*[@id="addOrderReceive"]/tbody/tr[2]/td[1]/div/a').click()
        waitxp(browser, '//*[@id="addOrderReceive"]/tbody/tr[2]/td[1]/div/ul/li[2]/a')
        browser.find_element_by_xpath('//*[@id="addOrderReceive"]/tbody/tr[2]/td[1]/div/ul/li[2]/a').click()
        browser.find_element_by_xpath('//*[@id="addOrderReceive"]/tbody/tr[2]/td[4]/div/input').send_keys(y)
        browser.find_element_by_id("submitBook").click()
        waitclass(browser,'help-block inline-help-block inline-help-block-right pull-right')
        str = browser.find_element_by_class_name('help-block inline-help-block inline-help-block-right pull-right').text
        print(str)



        # waitid(browser, "initCheckIn")
        # browser.find_element_by_id("initCheckIn").click()
        # waitid(browser, "checkInConfirm")
        # browser.find_element_by_id("checkInConfirm").click()
        # waitid(browser, "confirmEnterBtnL")
        # browser.find_element_by_id("confirmEnterBtnL").click()


if __name__ == "__main__":
    browser = webdriver.Firefox()
    login(browser)
    booking(browser)
